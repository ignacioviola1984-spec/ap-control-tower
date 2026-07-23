"""Maestro de proveedores Sage y resolucion segura de identidad.

La fuente se mantiene en memoria durante la sesion. No se persisten nombres,
tax IDs, cuentas ni el workbook. El matching prioriza Tax ID exacto; sin Tax ID
usa nombre fuertemente normalizado y solo despues similitud fuzzy.
"""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO
import re
import unicodedata
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..matching import FUZZY_SIMILARITY_THRESHOLD, similarity_score


FUZZY_VENDOR_FYI = (
    "proveedor vinculado por similitud de nombre, sin tax ID que lo confirme"
)
AMBIGUOUS_VENDOR_WARNING = (
    "proveedor ambiguo en maestro de Sage: múltiples candidatos posibles"
)
VENDOR_NOT_FOUND_WARNING = "proveedor no encontrado en maestro de Sage"
TAX_ID_NOT_FOUND_WARNING = (
    "tax ID del proveedor no encontrado en maestro de Sage"
)
MISSING_VENDOR_IDENTITY_WARNING = (
    "identidad del proveedor insuficiente para buscar en maestro de Sage"
)

MAX_WORKBOOK_BYTES = 20 * 1024 * 1024
MAX_VENDOR_ROWS = 20_000
_OPTIONAL_EMPTY_SENTINELS = {"168"}


class SageMasterError(ValueError):
    """Error de validacion seguro y accionable del export de Sage."""


@dataclass(frozen=True)
class SageVendor:
    source_id: str
    accounting_code: str | None
    legal_name: str
    trading_name: str | None
    tax_id_keys: tuple[str, ...]
    country_code: str | None
    iban: str | None
    bank_code: str | None
    payment_terms_code: str | None
    source_row: int
    active: bool = True

    @property
    def normalized_names(self) -> tuple[str, ...]:
        values = {
            normalize_supplier_name(self.legal_name),
            normalize_supplier_name(self.trading_name),
        }
        return tuple(sorted(value for value in values if value))


@dataclass(frozen=True)
class SageVendorMaster:
    vendors: tuple[SageVendor, ...]
    fingerprint: str
    source_filename: str
    sheet_name: str
    rows_seen: int
    rows_ignored: int
    inactive_count: int
    issues: tuple[str, ...] = ()

    @property
    def active_vendors(self) -> tuple[SageVendor, ...]:
        return tuple(vendor for vendor in self.vendors if vendor.active)

    def safe_summary(self) -> dict:
        return {
            "source": "sage-xlsx",
            "fingerprint": self.fingerprint,
            "rows_seen": self.rows_seen,
            "vendors": len(self.vendors),
            "active_vendors": len(self.active_vendors),
            "inactive_vendors": self.inactive_count,
            "rows_ignored": self.rows_ignored,
            "issues": len(self.issues),
        }


@dataclass(frozen=True)
class SupplierResolution:
    status: str
    method: str | None
    candidate_count: int
    score: float | None = None
    tax_id_confirmed: bool = False
    warning: str | None = None

    @property
    def accepted(self) -> bool:
        return self.status == "matched" and self.candidate_count == 1

    def safe_dict(self) -> dict:
        return {
            "status": self.status,
            "method": self.method,
            "candidate_count": self.candidate_count,
            "score": round(self.score, 4) if self.score is not None else None,
            "tax_id_confirmed": self.tax_id_confirmed,
            "warning": self.warning,
        }


_LEGAL_SUFFIXES = (
    ("sociedad", "de", "responsabilidad", "limitada"),
    ("sociedad", "limitada", "unipersonal"),
    ("sociedad", "anonima", "unipersonal"),
    ("sociedad", "limitada"),
    ("sociedad", "anonima"),
    ("s", "r", "l"),
    ("s", "l", "u"),
    ("s", "a", "u"),
    ("s", "a", "s"),
    ("s", "l"),
    ("s", "a"),
    ("l", "t", "d", "a"),
    ("l", "t", "d"),
    ("l", "l", "c"),
    ("l", "d", "a"),
    ("p", "l", "c"),
    ("g", "m", "b", "h"),
    ("b", "v"),
    ("n", "v"),
    ("srl",),
    ("slu",),
    ("sau",),
    ("sl",),
    ("sa",),
    ("sarl",),
    ("sas",),
    ("spa",),
    ("sro",),
    ("srls",),
    ("ltda",),
    ("lda",),
    ("limited",),
    ("ltd",),
    ("llc",),
    ("incorporated",),
    ("inc",),
    ("corporation",),
    ("corp",),
    ("company",),
    ("co",),
    ("gmbh",),
    ("ag",),
    ("bv",),
    ("nv",),
    ("plc",),
)


def _without_diacritics(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize_supplier_name(value) -> str:
    """Casefold, quita acentos/puntuacion y limpia sufijos legales finales."""
    text = _without_diacritics(str(value or "").casefold())
    tokens = re.sub(r"[^a-z0-9]+", " ", text).split()
    changed = True
    while tokens and changed:
        changed = False
        for suffix in _LEGAL_SUFFIXES:
            size = len(suffix)
            if len(tokens) >= size and tuple(tokens[-size:]) == suffix:
                del tokens[-size:]
                changed = True
                break
    return " ".join(tokens)


def normalize_tax_id(value) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _tax_keys(values: list[object], country_code: str | None) -> tuple[str, ...]:
    keys: set[str] = set()
    country = normalize_tax_id(country_code)
    for value in values:
        normalized = normalize_tax_id(value)
        if not normalized:
            continue
        keys.add(normalized)
        if len(country) == 2:
            if normalized.startswith(country) and len(normalized) > 2:
                keys.add(normalized[2:])
            else:
                keys.add(country + normalized)
    return tuple(sorted(keys))


def _normalize_header(value) -> str:
    text = _without_diacritics(str(value or "").casefold())
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


_HEADER_ALIASES = {
    "provider_code": ("cod proveedor", "codigo proveedor"),
    "client_code": ("cod cliente", "codigo cliente"),
    "client_category": ("cod categoria cliente", "codigo categoria cliente"),
    "accounting_code": ("cod contable", "codigo contable"),
    "legal_name": ("razon social",),
    "trading_name": ("nombre cli pro", "nombre proveedor", "nombre comercial"),
    "tax_id": ("cif dni", "nif cif", "tax id", "id fiscal"),
    "eu_tax_id": ("cif europeo", "vat id"),
    "country": ("sigla nacion", "codigo pais", "pais"),
    "iban": ("i b a n", "iban"),
    "bank_code": ("cod banco", "codigo banco"),
    "payment_terms": ("cod condiciones", "codigo condiciones", "condiciones pago"),
    "inactive": ("baja empresa", "inactivo"),
    "inactive_date": ("fecha baja",),
}


def _column_map(headers: tuple[object, ...]) -> dict[str, int]:
    normalized: dict[str, int] = {}
    for index, value in enumerate(headers):
        normalized.setdefault(_normalize_header(value), index)
    result: dict[str, int] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[field] = normalized[alias]
                break
    return result


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _value(row: tuple, columns: dict[str, int], field: str) -> str:
    index = columns.get(field)
    return _text(row[index]) if index is not None and index < len(row) else ""


def _optional_value(row: tuple, columns: dict[str, int], field: str) -> str:
    value = _value(row, columns, field)
    return "" if value in _OPTIONAL_EMPTY_SENTINELS else value


def _find_header(rows: list[tuple]) -> tuple[int, tuple, dict[str, int]]:
    for index, row in enumerate(rows[:20]):
        columns = _column_map(row)
        if "legal_name" in columns and (
            "provider_code" in columns or "accounting_code" in columns
        ):
            return index, row, columns
    raise SageMasterError(
        "No se encontraron las columnas requeridas de Sage: Cód. proveedor y Razón social."
    )


def _looks_like_customer_export(data_rows: list[tuple], columns: dict[str, int]) -> bool:
    if not {"provider_code", "client_code"} <= columns.keys() or len(data_rows) < 3:
        return False
    provider_codes = {
        _value(row, columns, "provider_code") for row in data_rows
        if _value(row, columns, "provider_code")
    }
    client_codes = {
        _value(row, columns, "client_code") for row in data_rows
        if _value(row, columns, "client_code")
    }
    categories = [
        _value(row, columns, "client_category").upper() for row in data_rows
        if _value(row, columns, "client_category")
    ]
    client_category_ratio = (
        sum(value == "CLI" for value in categories) / len(categories)
        if categories else 0.0
    )
    return (
        len(provider_codes) <= 1
        and len(client_codes) >= 3
        and (client_category_ratio >= 0.8 or len(client_codes) > len(provider_codes) * 3)
    )


def _is_inactive(row: tuple, columns: dict[str, int]) -> bool:
    flag = _value(row, columns, "inactive").casefold()
    if flag in {"si", "sí", "yes", "true", "1", "baja", "inactivo"}:
        return True
    return bool(_optional_value(row, columns, "inactive_date"))


def load_vendor_master_xlsx(
    content: bytes, *, filename: str = "maestro-proveedores-sage.xlsx"
) -> SageVendorMaster:
    """Carga un export XLSX de proveedores y descarta el workbook al retornar."""
    if not content:
        raise SageMasterError("El archivo de Sage está vacío.")
    if len(content) > MAX_WORKBOOK_BYTES:
        raise SageMasterError("El archivo de Sage supera el límite permitido de 20 MB.")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError):
        raise SageMasterError("El archivo no es un XLSX válido de Sage.") from None
    try:
        if not workbook.sheetnames:
            raise SageMasterError("El archivo de Sage no contiene hojas.")
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True, max_row=MAX_VENDOR_ROWS + 21))
        header_index, _, columns = _find_header(rows)
        data_rows = [row for row in rows[header_index + 1:] if any(value is not None for value in row)]
        if len(data_rows) > MAX_VENDOR_ROWS:
            raise SageMasterError(
                f"El maestro supera el límite operativo de {MAX_VENDOR_ROWS:,} filas."
            )
        if _looks_like_customer_export(data_rows, columns):
            raise SageMasterError(
                "El archivo corresponde a un maestro de clientes, no de proveedores: "
                "Cód. cliente varía y Cód. proveedor no. Exportá el maestro de proveedores de Sage."
            )

        vendors: list[SageVendor] = []
        issues: list[str] = []
        ignored = inactive_count = 0
        for offset, row in enumerate(data_rows, start=header_index + 2):
            provider_code = _value(row, columns, "provider_code")
            accounting_code = _optional_value(row, columns, "accounting_code") or None
            source_id = provider_code or accounting_code
            legal_name = _optional_value(row, columns, "legal_name")
            trading_name = _optional_value(row, columns, "trading_name") or None
            if not source_id:
                ignored += 1
                issues.append(f"fila {offset}: sin código de proveedor")
                continue
            if not legal_name and not trading_name:
                ignored += 1
                issues.append(f"fila {offset}: sin nombre de proveedor")
                continue
            country = _optional_value(row, columns, "country").upper() or None
            inactive = _is_inactive(row, columns)
            inactive_count += int(inactive)
            vendors.append(SageVendor(
                source_id=source_id,
                accounting_code=accounting_code,
                legal_name=legal_name or str(trading_name),
                trading_name=trading_name,
                tax_id_keys=_tax_keys([
                    _optional_value(row, columns, "tax_id"),
                    _optional_value(row, columns, "eu_tax_id"),
                ], country),
                country_code=country,
                iban=_optional_value(row, columns, "iban") or None,
                bank_code=_optional_value(row, columns, "bank_code") or None,
                payment_terms_code=_optional_value(row, columns, "payment_terms") or None,
                source_row=offset,
                active=not inactive,
            ))
        if not vendors:
            raise SageMasterError("El archivo no contiene proveedores utilizables.")
        if not any(vendor.active for vendor in vendors):
            raise SageMasterError("El archivo no contiene proveedores activos.")
        return SageVendorMaster(
            vendors=tuple(vendors),
            fingerprint=sha256(content).hexdigest()[:16],
            source_filename=filename,
            sheet_name=sheet.title,
            rows_seen=len(data_rows),
            rows_ignored=ignored,
            inactive_count=inactive_count,
            issues=tuple(issues),
        )
    finally:
        workbook.close()


def _document_names(document: dict) -> tuple[str, ...]:
    names = {
        normalize_supplier_name(document.get("proveedor_razon_social_legal")),
        normalize_supplier_name(document.get("proveedor_nombre_comercial")),
    }
    return tuple(sorted(name for name in names if name))


def resolve_document_supplier(
    document: dict, master: SageVendorMaster
) -> SupplierResolution:
    """Resuelve proveedor con prioridad Tax ID > exacto normalizado > fuzzy."""
    vendors = master.active_vendors
    tax_id = normalize_tax_id(document.get("proveedor_tax_id"))
    if tax_id:
        candidates = [vendor for vendor in vendors if tax_id in vendor.tax_id_keys]
        if len(candidates) == 1:
            return SupplierResolution(
                "matched", "tax_id", 1, score=1.0, tax_id_confirmed=True
            )
        if len(candidates) > 1:
            return SupplierResolution(
                "ambiguous", "tax_id", len(candidates), score=1.0,
                tax_id_confirmed=True, warning=AMBIGUOUS_VENDOR_WARNING,
            )
        return SupplierResolution(
            "not_found", "tax_id", 0, tax_id_confirmed=False,
            warning=TAX_ID_NOT_FOUND_WARNING,
        )

    names = _document_names(document)
    if not names:
        return SupplierResolution(
            "not_found", None, 0, warning=MISSING_VENDOR_IDENTITY_WARNING
        )

    exact = [
        vendor for vendor in vendors
        if any(name in vendor.normalized_names for name in names)
    ]
    if len(exact) == 1:
        return SupplierResolution("matched", "exact_name", 1, score=1.0)
    if len(exact) > 1:
        return SupplierResolution(
            "ambiguous", "exact_name", len(exact), score=1.0,
            warning=AMBIGUOUS_VENDOR_WARNING,
        )

    scored: list[tuple[SageVendor, float]] = []
    for vendor in vendors:
        score = max(
            (similarity_score(name, candidate)
             for name in names for candidate in vendor.normalized_names),
            default=0.0,
        )
        if score >= FUZZY_SIMILARITY_THRESHOLD:
            scored.append((vendor, score))
    if len(scored) == 1:
        return SupplierResolution(
            "matched", "fuzzy_name", 1, score=scored[0][1],
            warning=FUZZY_VENDOR_FYI,
        )
    if len(scored) > 1:
        return SupplierResolution(
            "ambiguous", "fuzzy_name", len(scored),
            score=max(score for _, score in scored),
            warning=AMBIGUOUS_VENDOR_WARNING,
        )
    return SupplierResolution(
        "not_found", "fuzzy_name", 0, warning=VENDOR_NOT_FOUND_WARNING
    )
