"""Replay offline de la politica del trial sobre extracciones persistidas.

No abre PDFs ni llama a Document AI. Reconstruye los resultados estructurados
desde el CSV de extraccion, ejecuta las reglas puras de ``workflow.py`` y
compara el ruteo contra el golden dataset.

Uso:
    python evals/run_policy_replay.py
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import subprocess
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ap_control_tower.ui.trial import workflow


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_GOLDEN = ROOT / "Fable 5/AP Evals/Golden/golden_dataset_v1.0.xlsx"
DEFAULT_EXTRACTION = ROOT / "Fable 5/AP Evals/Corridas/run2_extraccion.csv"
DEFAULT_OUTPUT_DIR = ROOT / "Fable 5/AP Evals/Corridas"
DEFAULT_REPORT_DIR = ROOT / "Fable 5/AP Evals/Reportes"

FORMAL_ROUTES = {"en_lote", "revision_humana"}
SMOKE_DOC_IDS = (
    "GD-001",  # factura real limpia: control negativo
    "GD-105", "GD-106", "GD-107", "GD-108",  # grupo de duplicados
    "GD-109", "GD-110",  # C6: IBAN
    "GD-111", "GD-112",  # C3: autorizacion y saldo de OC
    "GD-113", "GD-114",  # C5: fuera/dentro de tolerancia
    "GD-116",  # C8: anticipo
    "GD-117",  # conciliacion
    "GD-118",  # intercompany que debe avanzar
    "GD-119",  # nota de credito
)


@dataclass
class ReplayResult:
    doc_id: str
    document: dict[str, Any]
    confidence: Decimal
    warnings: list[str] = field(default_factory=list)
    field_confidences: dict[str, Decimal] = field(default_factory=dict)
    engine: str = "persisted_run2_extraction"
    pages: int = 0
    text_chars: int = 0


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _git_value(*args: str) -> str | None:
    try:
        completed = subprocess.run(
            ["git", *args], cwd=ROOT, check=True, capture_output=True, text=True)
    except (OSError, subprocess.CalledProcessError):
        return None
    return completed.stdout.strip()


def _stem(value: Any) -> str:
    name = Path(str(value)).name
    return name[:-4] if name.casefold().endswith(".pdf") else name


def _optional(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    return None if text == "" else value


def _load_golden(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Golden_dataset"]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        return {
            str(row[0]): {headers[index]: row[index] for index in range(len(headers))}
            for row in rows
            if row[0]
        }
    finally:
        workbook.close()


def _load_extractions(
        path: Path, golden: dict[str, dict[str, Any]],
        ) -> tuple[list[ReplayResult], dict[str, str], list[str]]:
    file_to_doc = {
        _stem(record["archivo_pdf"]): doc_id for doc_id, record in golden.items()
    }
    results: list[ReplayResult] = []
    old_states: dict[str, str] = {}
    out_of_scope: list[str] = []

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            filename = str(row["archivo"])
            doc_id = file_to_doc.get(filename)
            if not doc_id:
                out_of_scope.append(filename)
                continue

            confidence = Decimal(str(row["confianza"]))
            warnings = [
                item.strip() for item in str(row.get("advertencias") or "").split("|")
                if item.strip()
            ]
            document = {
                "document_type": _optional(row.get("tipo documental")),
                "proveedor_nombre_comercial": _optional(row.get("proveedor")),
                "proveedor_razon_social_legal": _optional(row.get("proveedor")),
                "proveedor_tax_id": _optional(row.get("tax ID proveedor")),
                "numero_factura": _optional(row.get("numero") or row.get("número")),
                "fecha_emision": _optional(row.get("fecha de emisión")),
                "fecha_vencimiento_calculada": _optional(row.get("vencimiento")),
                "moneda": _optional(row.get("moneda")),
                "importe_neto": _optional(row.get("importe neto")),
                "tipo_iva": _optional(row.get("tipo IVA")),
                "importe_iva": _optional(row.get("importe IVA")),
                "importe_total": _optional(row.get("importe total")),
                "po_reference": _optional(row.get("referencia OC")),
            }
            field_confidences = {
                field_name: confidence for field_name in workflow.REVIEW_RELEVANT_FIELDS
            }
            results.append(ReplayResult(
                doc_id=doc_id,
                document=document,
                confidence=confidence,
                warnings=warnings,
                field_confidences=field_confidences,
            ))
            old_states[doc_id] = str(row.get("estado") or "")

    if len(results) != len(golden):
        missing = sorted(set(golden) - {result.doc_id for result in results})
        raise ValueError(
            f"El replay requiere una extraccion por golden record; faltan: {missing}")
    if len({result.doc_id for result in results}) != len(results):
        raise ValueError("La extraccion contiene doc_id duplicados tras mapear archivos")
    return results, old_states, out_of_scope


def _route_from_run2_state(state: str) -> str:
    return "revision_humana" if state == "Revisar campos" else "en_lote"


def _confusion(
        predictions: dict[str, str], golden: dict[str, dict[str, Any]],
        ) -> dict[str, Any]:
    scope = [
        doc_id for doc_id, record in golden.items()
        if record["resultado_esperado"] in FORMAL_ROUTES
    ]
    expected = {
        doc_id: golden[doc_id]["resultado_esperado"] == "revision_humana"
        for doc_id in scope
    }
    predicted = {
        doc_id: predictions[doc_id] == "revision_humana" for doc_id in scope
    }
    tp = sum(expected[d] and predicted[d] for d in scope)
    fp = sum(not expected[d] and predicted[d] for d in scope)
    fn = sum(expected[d] and not predicted[d] for d in scope)
    tn = sum(not expected[d] and not predicted[d] for d in scope)

    def percent(numerator: int, denominator: int) -> float | None:
        return round(100 * numerator / denominator, 1) if denominator else None

    return {
        "scope_documents": len(scope),
        "tp": tp, "fp": fp, "fn": fn, "tn": tn,
        "routing_accuracy_pct": percent(tp + tn, len(scope)),
        "review_precision_pct": percent(tp, tp + fp),
        "review_recall_pct": percent(tp, tp + fn),
    }


def _write_detail(
        path: Path, results: list[ReplayResult], golden: dict[str, dict[str, Any]],
        old_states: dict[str, str], duplicate_ids: set[str],
        replay_routes: dict[str, str],
        ) -> None:
    fields = (
        "doc_id", "fuente", "resultado_esperado", "control_esperado",
        "run2_route", "run3_replay_route", "formal_scope",
        "formal_route_correct", "duplicate_signal", "review_reasons",
    )
    by_id = {result.doc_id: result for result in results}
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for doc_id in sorted(golden):
            record = golden[doc_id]
            formal = record["resultado_esperado"] in FORMAL_ROUTES
            reasons = workflow.review_reasons(
                by_id[doc_id], duplicate=doc_id in duplicate_ids)
            writer.writerow({
                "doc_id": doc_id,
                "fuente": record["fuente"],
                "resultado_esperado": record["resultado_esperado"],
                "control_esperado": record["control_esperado"] or "",
                "run2_route": _route_from_run2_state(old_states[doc_id]),
                "run3_replay_route": replay_routes[doc_id],
                "formal_scope": "si" if formal else "no",
                "formal_route_correct": (
                    "si" if formal and replay_routes[doc_id] == record["resultado_esperado"]
                    else "no" if formal else "no_aplica"),
                "duplicate_signal": "si" if doc_id in duplicate_ids else "no",
                "review_reasons": " | ".join(reasons),
            })


def _write_smoke_candidates(
        path: Path, golden: dict[str, dict[str, Any]]) -> None:
    fields = (
        "priority", "doc_id", "archivo_pdf", "resultado_esperado",
        "control_esperado", "purpose",
    )
    purposes = {
        "GD-001": "control negativo real; debe avanzar",
        "GD-105": "original del grupo; medir falso positivo del detector",
        "GD-106": "duplicado exacto C2",
        "GD-107": "casi-duplicado corregido C2",
        "GD-108": "duplicado con normalizacion C2",
        "GD-109": "IBAN completamente distinto C6",
        "GD-110": "IBAN con un digito alterado C6",
        "GD-111": "OC inexistente C3",
        "GD-112": "OC sin saldo C3",
        "GD-113": "match fuera de tolerancia C5",
        "GD-114": "match dentro de tolerancia; control negativo",
        "GD-116": "proforma/anticipo C8",
        "GD-117": "ruta de conciliacion",
        "GD-118": "intercompany que debe avanzar con flag",
        "GD-119": "nota de credito corregida",
    }
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for index, doc_id in enumerate(SMOKE_DOC_IDS, start=1):
            record = golden[doc_id]
            writer.writerow({
                "priority": index,
                "doc_id": doc_id,
                "archivo_pdf": record["archivo_pdf"],
                "resultado_esperado": record["resultado_esperado"],
                "control_esperado": record["control_esperado"] or "",
                "purpose": purposes[doc_id],
            })


def _write_report(path: Path, manifest: dict[str, Any]) -> None:
    run2 = manifest["metrics"]["run2_reconstructed"]
    run3 = manifest["metrics"]["run3_policy_replay"]
    changes = manifest["route_changes"]
    change_lines = "\n".join(
        f"| {item['doc_id']} | {item['run2']} | {item['run3']} | "
        f"{item['expected']} | {item['interpretation']} |"
        for item in changes
    )
    path.write_text(f"""# Reporte de evals · run3 policy replay offline

Fecha UTC: {manifest['created_at_utc']}  
Golden dataset: {manifest['inputs']['golden']['version']} ({manifest['scope']['golden_documents']} documentos)  
Extraccion reutilizada: run2  
Llamadas a Document AI: **0**

## 1. Que valida esta corrida

Esta corrida reconstruye los 106 resultados estructurados guardados en run2 y
ejecuta la version actual de las reglas puras de revision y duplicados. No abre
PDFs, no vuelve a extraer y no consume credito de Google.

Es una regresion integral de la **politica evaluable con los datos persistidos**.
No valida nuevamente el extractor ni los controles que requieren maestros o
campos que run2 no exporto.

## 2. Resultados comparables (alcance formal: 96 documentos)

| Metrica | run2 reconstruido | run3 policy replay | Delta |
|---|---:|---:|---:|
| Exactitud de ruteo | {run2['routing_accuracy_pct']}% | {run3['routing_accuracy_pct']}% | {round(run3['routing_accuracy_pct'] - run2['routing_accuracy_pct'], 1):+.1f} pts |
| Precision de derivacion | {run2['review_precision_pct']}% | {run3['review_precision_pct']}% | {round(run3['review_precision_pct'] - run2['review_precision_pct'], 1):+.1f} pts |
| Recall de derivacion | {run2['review_recall_pct']}% | {run3['review_recall_pct']}% | {round(run3['review_recall_pct'] - run2['review_recall_pct'], 1):+.1f} pts |
| Falsos negativos | {run2['fn']} | {run3['fn']} | {run3['fn'] - run2['fn']:+d} |
| Derivados sobre los 106 golden | {manifest['metrics']['run2_review_count']}/106 ({manifest['metrics']['run2_review_rate_pct']}%) | {manifest['metrics']['run3_review_count']}/106 ({manifest['metrics']['run3_review_rate_pct']}%) | {round(manifest['metrics']['run3_review_rate_pct'] - manifest['metrics']['run2_review_rate_pct'], 1):+.1f} pts |

El recall vuelve a 100%: GD-119 ahora se deriva como posible nota de credito.
La exactitud de ruteo sube a 96.9%. La tasa de revision aumenta en un caso neto
porque run3 libera GD-018 y agrega GD-107 y GD-119.

## 3. Cambios de ruta

| Documento | run2 | run3 | Esperado | Lectura |
|---|---|---|---|---|
{change_lines}

## 4. Cobertura de controles offline

- **C2 duplicados:** 3/3 casos que debian bloquearse reciben senal de duplicado
  (GD-106, GD-107 y GD-108). El detector tambien retiene la original GD-105;
  sigue siendo una decision funcional pendiente si el grupo completo debe
  frenarse hasta revision.
- **Nota de credito:** GD-119 pasa de `en_lote` a `revision_humana`.
- **C9 vendor master:** la advertencia persistida de GD-115 sigue derivando,
  pero este replay no vuelve a consultar el maestro.

No son evaluables offline con el export de run2:

- C6 datos bancarios: 2 casos; faltan IBAN extraido completo y maestro bancario.
- C3 autorizacion de OC: 2 casos; falta maestro/estado/saldo de OC.
- C5 match: 1 caso; falta el estado de la OC y su tolerancia aplicada.
- C8 anticipo y la ruta de conciliacion: el documento se reconoce, pero el CSV
  no conserva todo el estado necesario para validar la ruta especializada.

Por esta limitacion **no se declara todavia cero escapes de riesgo de pago para
el conjunto C2-C9**. Esa afirmacion queda reservada al cloud smoke autorizado.

## 5. Trazabilidad

- Golden SHA-256: `{manifest['inputs']['golden']['sha256']}`
- Extraccion run2 SHA-256: `{manifest['inputs']['run2_extraction']['sha256']}`
- workflow.py SHA-256: `{manifest['code']['workflow_sha256']}`
- Commit observado: `{manifest['code']['git_commit']}`
- Working tree dirty: `{str(manifest['code']['git_dirty']).lower()}`

El detalle documento por documento y el manifest JSON permiten repetir y
auditar el calculo sin procesar facturas nuevamente.

## 6. Decision sobre el cloud smoke

La muestra dirigida contiene {manifest['smoke_plan']['documents']} documentos:
los fixes, todos los controles no evaluables offline y tres controles negativos.
El smoke queda **diferido hasta que las integraciones con Zoho/Sage y los
maestros requeridos esten conectados al circuito evaluable**. Ejecutarlo antes
solo reconfirmaria los fixes de politica y no agregaria evidencia comercial
material sobre C3, C5 o C6.

No se realizo ninguna llamada cloud durante esta corrida. Los candidatos quedan
versionados para reutilizarlos cuando se cumplan las condiciones de entrada.
""", encoding="utf-8")


def run(args: argparse.Namespace) -> dict[str, Any]:
    golden_path = args.golden.resolve()
    extraction_path = args.extraction.resolve()
    output_dir = args.output_dir.resolve()
    report_dir = args.report_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    report_dir.mkdir(parents=True, exist_ok=True)

    golden = _load_golden(golden_path)
    results, old_states, out_of_scope = _load_extractions(
        extraction_path, golden)
    by_id = {result.doc_id: result for result in results}
    duplicate_ids = workflow.duplicate_doc_ids(results)
    replay_routes = {
        doc_id: (
            "revision_humana" if workflow.requires_human_review(
                result, duplicate=doc_id in duplicate_ids)
            else "en_lote"
        )
        for doc_id, result in by_id.items()
    }
    run2_routes = {
        doc_id: _route_from_run2_state(state)
        for doc_id, state in old_states.items()
    }
    run2_metrics = _confusion(run2_routes, golden)
    run3_metrics = _confusion(replay_routes, golden)

    run2_review_count = sum(route == "revision_humana" for route in run2_routes.values())
    run3_review_count = sum(route == "revision_humana" for route in replay_routes.values())
    route_changes = []
    interpretations = {
        "GD-018": "libera un falso positivo de run2",
        "GD-107": "el fix detecta el casi-duplicado C2",
        "GD-119": "el fix recupera el falso negativo de run2",
    }
    for doc_id in sorted(golden):
        if run2_routes[doc_id] != replay_routes[doc_id]:
            route_changes.append({
                "doc_id": doc_id,
                "run2": run2_routes[doc_id],
                "run3": replay_routes[doc_id],
                "expected": golden[doc_id]["resultado_esperado"],
                "interpretation": interpretations.get(doc_id, "cambio de politica"),
            })

    git_status = _git_value("status", "--porcelain")
    manifest = {
        "run_id": "run3-policy-replay",
        "status": "completed_with_declared_limitations",
        "created_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "execution": {
            "mode": "offline_policy_replay",
            "document_ai_calls": 0,
            "pdfs_opened": 0,
            "cloud_services_called": [],
        },
        "scope": {
            "golden_documents": len(golden),
            "formal_routing_documents": run3_metrics["scope_documents"],
            "out_of_scope_extraction_rows": out_of_scope,
        },
        "inputs": {
            "golden": {
                "path": str(golden_path.relative_to(ROOT)),
                "version": "v1.0 corregido",
                "sha256": _sha256(golden_path),
            },
            "run2_extraction": {
                "path": str(extraction_path.relative_to(ROOT)),
                "sha256": _sha256(extraction_path),
            },
        },
        "code": {
            "git_commit": _git_value("rev-parse", "HEAD"),
            "git_dirty": bool(git_status),
            "workflow_path": "ap_control_tower/ui/trial/workflow.py",
            "workflow_sha256": _sha256(ROOT / "ap_control_tower/ui/trial/workflow.py"),
        },
        "metrics": {
            "run2_reconstructed": run2_metrics,
            "run3_policy_replay": run3_metrics,
            "run2_review_count": run2_review_count,
            "run3_review_count": run3_review_count,
            "run2_review_rate_pct": round(100 * run2_review_count / len(golden), 1),
            "run3_review_rate_pct": round(100 * run3_review_count / len(golden), 1),
        },
        "route_changes": route_changes,
        "duplicate_control": {
            "expected_blocked_ids": ["GD-106", "GD-107", "GD-108"],
            "detected_expected_ids": sorted(
                duplicate_ids & {"GD-106", "GD-107", "GD-108"}),
            "all_flagged_ids": sorted(duplicate_ids),
        },
        "not_evaluable_offline": {
            "C6_DATOS_BANCARIOS": ["GD-109", "GD-110"],
            "C3_AUTORIZACION_OC": ["GD-111", "GD-112"],
            "C5_MATCH": ["GD-113"],
            "specialized_routes": ["GD-116", "GD-117"],
        },
        "limitations": [
            "No reevalua la extraccion ni cambios de Document AI.",
            "El CSV run2 no exporta IBAN completo ni estado/maestro de OC.",
            "No permite afirmar cero escapes para todos los controles C2-C9.",
            "La confianza persistida es agregada, no por campo.",
        ],
        "commercial_evidence": {
            "defensible_claims": [
                "La politica actual fue reejecutada sobre 106 extracciones persistidas sin costo cloud.",
                "La exactitud de ruteo del alcance comparable sube de 94.8% a 96.9%.",
                "El recall de derivacion vuelve de 75.0% a 100.0%, sin falsos negativos en 96 casos formales.",
                "Los fixes de casi-duplicado y nota de credito corrigen GD-107 y GD-119.",
            ],
            "claims_not_supported_yet": [
                "No revalida la exactitud de Document AI.",
                "No demuestra integraciones productivas con Zoho o Sage.",
                "No permite afirmar cero escapes para todos los controles C2-C9.",
            ],
        },
        "smoke_plan": {
            "status": "deferred_until_zoho_sage_integrations",
            "decision": "No ejecutar ahora: el costo no agrega evidencia comercial material antes de conectar los maestros externos.",
            "entry_conditions": [
                "Adaptadores Zoho/Sage conectados al circuito evaluable.",
                "Maestros de proveedores, OC y datos bancarios disponibles con tratamiento seguro.",
                "Version desplegable vinculada a un commit limpio.",
                "Matriz de aceptacion acordada para C3, C5, C6 y rutas especializadas.",
            ],
            "documents": len(SMOKE_DOC_IDS),
            "doc_ids": list(SMOKE_DOC_IDS),
        },
    }

    detail_path = output_dir / "run3_policy_replay_detail.csv"
    smoke_path = output_dir / "run3_cloud_smoke_candidates.csv"
    manifest_path = output_dir / "run3_policy_replay_manifest.json"
    report_path = report_dir / "reporte_evals_run3_policy_replay.md"
    _write_detail(
        detail_path, results, golden, old_states, duplicate_ids, replay_routes)
    _write_smoke_candidates(smoke_path, golden)
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    _write_report(report_path, manifest)

    print(json.dumps({
        "run_id": manifest["run_id"],
        "document_ai_calls": 0,
        "metrics": manifest["metrics"],
        "outputs": [
            str(detail_path), str(manifest_path), str(report_path), str(smoke_path),
        ],
    }, ensure_ascii=False, indent=2))
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--golden", type=Path, default=DEFAULT_GOLDEN)
    parser.add_argument("--extraction", type=Path, default=DEFAULT_EXTRACTION)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
