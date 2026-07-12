"""Eval: estado activo del trial y privacidad. exit 0 = verde.

Hermetico (no arranca Streamlit, no toca red ni disco): usa un resultado de
extraccion de mentira (duck-typed) para ejercitar el modelo puro de la sesion y
los helpers de presentacion.

Valida: los resultados viven en la sesion; el audit trail es temporal y sin PII;
'Finalizar y borrar' limpia las claves de la sesion; una sesion nueva no ve
resultados anteriores (aislamiento).
"""

from __future__ import annotations

import sys
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

failures: list[str] = []


def check(cond: bool, label: str) -> None:
    print(f"  {'PASS' if cond else 'FAIL'}  {label}")
    if not cond:
        failures.append(label)


@dataclass
class FakeResult:
    doc_id: str
    document: dict
    engine: str = "fallback_local"
    confidence: Decimal = Decimal("0.42")
    pages: int = 1
    text_chars: int = 120
    warnings: list = field(default_factory=list)
    field_confidences: dict = field(default_factory=dict)


def _fake_result(doc_id: str, numero: str, total: str, warn: bool) -> FakeResult:
    from ap_control_tower.extraction.schema import empty_document
    doc = empty_document()
    doc["document_type"] = "invoice"
    doc["numero_factura"] = numero
    doc["fecha_emision"] = "2026-05-01"
    doc["proveedor_nombre_comercial"] = "Proveedor Test SL"
    doc["proveedor_tax_id"] = "ESB12345678"
    doc["importe_total"] = total
    doc["moneda"] = "EUR"
    return FakeResult(
        doc_id=doc_id, document=doc,
        warnings=(["campos criticos ausentes"] if warn else []),
        field_confidences={"numero_factura": Decimal("0.9")},
    )


def main() -> int:
    from ap_control_tower.ui.components import extraction_view as ev
    from ap_control_tower.ui.trial import session as se

    print("== Modelo de sesion: resultados + audit temporal ==")
    s = se.new_session()
    check([e.action for e in s.audit.events] == ["sesion-iniciada"],
          "la sesion arranca con un evento 'sesion-iniciada'")
    r1 = _fake_result("DOC-1", "F-001", "121.00", warn=False)
    r2 = _fake_result("DOC-2", "F-002", "242.00", warn=True)
    se.add_results(s, [r1, r2])
    se.record_intake(s, canal="carga-manual", cantidad=2)
    check(len(s.results) == 2, "dos resultados guardados en la sesion")
    check(s.audit.verify_chain(), "cadena de auditoria de la sesion integra")
    actions = [e.action for e in s.audit.events]
    check(actions == ["sesion-iniciada", "documento-procesado", "documento-procesado", "ingesta"],
          f"eventos esperados en orden ({actions})")

    print("== Idempotencia: una carga repetida no contamina la corrida ==")
    duplicate_state = se.new_session()
    first = se.add_document(
        duplicate_state, r1, 0.5, file_hash="d" * 64, source="correo-ap")
    second = se.add_document(
        duplicate_state, r1, 0.5, file_hash="d" * 64, source="correo-ap")
    check(first is True and second is False and len(duplicate_state.results) == 1,
          "el mismo PDF se guarda una sola vez")
    check(duplicate_state.processing_seconds == 0.5,
          "la repetición omitida no duplica el tiempo")
    check(duplicate_state.audit.events[-1].action == "documento-repetido-omitido",
          "la omisión queda auditada sin reprocesar")

    legacy_state = se.new_session()
    legacy_state.results = [r1, r2, r1, r2]
    legacy_state.proc_seconds = {"DOC-1": 0.4, "DOC-2": 0.6}
    legacy_state.processing_seconds = 2.0
    removed = se.repair_duplicates(legacy_state)
    check(removed == 2 and len(legacy_state.results) == 2,
          "una sesión ya contaminada se repara automáticamente")
    check(legacy_state.processing_seconds == 1.0,
          "la reparación corrige el tiempo acumulado")
    check(ev.aggregate_metrics([r1, r2, r1, r2])["documents"] == 2,
          "métricas defensivas cuentan documentos únicos")
    check(len(ev.results_csv([r1, r2, r1, r2]).splitlines()) == 3,
          "exportación defensiva contiene una fila por documento único")

    print("== Privacidad: el audit trail no guarda contenido del documento ==")
    blob = " ".join(str(e.evidence) for e in s.audit.events)
    check("F-001" not in blob and "ESB12345678" not in blob and "121.00" not in blob,
          "ningun valor de campo/PII en la evidencia de auditoria")
    ev0 = s.audit.events[1].evidence
    check("tipo" in ev0 and "confianza" in ev0 and "motor" in ev0,
          "la evidencia guarda solo metadatos (tipo/confianza/motor)")

    print("== Helpers de presentacion (cobertura, CSV y Excel) ==")
    check(0.0 < ev.coverage(r1) < 1.0, "cobertura entre 0 y 1")
    check(len(ev.present_fields(r1)) + len(ev.missing_fields(r1)) == len(ev.BUSINESS_FIELDS),
          "encontrados + ausentes == total de campos de negocio")
    csv = ev.results_csv([r1, r2])
    check(csv.splitlines()[0].startswith("archivo,tipo documental,proveedor"),
          "CSV con encabezado esperado")
    check(len(csv.splitlines()) == 3, "CSV con una fila por documento")
    check("ESB12345678" not in csv,
          "CSV no revela tax ID completo")
    from io import BytesIO
    from openpyxl import load_workbook
    workbook = load_workbook(BytesIO(ev.results_excel([r1, r2])))
    sheet = workbook["Extracción"]
    excel_headers = [cell.value for cell in sheet[1]]
    csv_headers = csv.splitlines()[0].split(",")
    check(excel_headers == csv_headers == ev.EXPORT_COLUMNS,
          "CSV y Excel comparten columnas comerciales")
    check(sheet.max_row == 3, "Excel con una fila por documento")
    check("ESB12345678" not in " ".join(
        str(cell.value or "") for row in sheet.iter_rows() for cell in row),
        "Excel no revela tax ID completo")
    total_col = excel_headers.index("importe total") + 1
    check(isinstance(sheet.cell(2, total_col).value, (int, float)),
          "Excel conserva importes como números")
    from datetime import datetime
    date_col = excel_headers.index("fecha de emisión") + 1
    check(isinstance(sheet.cell(2, date_col).value, datetime),
          "Excel conserva fechas como fechas")

    print("== Descargas: claves unicas para sesion e historial ==")
    download_keys: list[str] = []
    original_columns = ev.st.columns
    class FakeColumn:
        def download_button(self, *args, **kwargs):
            download_keys.append(kwargs["key"])
    try:
        ev.st.columns = lambda count: [FakeColumn() for _ in range(count)]
        ev.render_download([r1], key="trial_download_current_run-1")
        ev.render_download([r1], key="trial_download_history_run-2")
    finally:
        ev.st.columns = original_columns
    check(download_keys == [
        "trial_download_current_run-1_csv", "trial_download_current_run-1_xlsx",
        "trial_download_history_run-2_csv", "trial_download_history_run-2_xlsx"],
        "sesión/historial y CSV/Excel usan IDs distintos")

    print("== 'Finalizar y borrar' limpia la sesion ==")
    to_clear = se.session_keys_to_clear(
        ["_trial_session", "_trial_uploader", "_gmail_demo_results", "otra_clave"])
    check("_trial_session" in to_clear and "_trial_uploader" in to_clear,
          "borra las claves de la sesion trial")
    check("otra_clave" not in to_clear, "no toca claves ajenas a la sesion trial")

    print("== Aislamiento: una sesion nueva no ve resultados anteriores ==")
    s2 = se.new_session()
    check(len(s2.results) == 0 and s2 is not s,
          "sesion nueva vacia e independiente de la anterior")

    print("== Estados, ruta PO/non-PO y confianza informada ==")
    r_ok = _fake_result("OK-1", "F-9", "100.00", warn=False)
    r_rev = _fake_result("REV-1", "F-10", "100.00", warn=True)
    check(ev.status_label(r_ok) == "Procesado", "sin advertencias -> 'Procesado'")
    check(ev.status_label(r_rev) == "Revisar campos", "advertencia de campo -> 'Revisar campos'")
    r_mode = _fake_result("MODE-1", "F-11", "100.00", warn=False)
    r_mode.warnings = ["Document AI no configurado; requiere revision"]
    check(ev.status_label(r_mode) == "Procesado",
          "nota de modo (Document AI) NO marca 'Revisar campos'")
    r_other = _fake_result("OTH-1", "", "0", warn=False)
    r_other.document["document_type"] = "other"
    check(ev.status_label(r_other) == "Documento no reconocido",
          "document_type 'other' -> 'Documento no reconocido'")
    check(ev.route_label(r_ok.document) == "non-PO" and ev.status_label(r_ok) == "Procesado",
          "sin OC -> ruta non-PO normal, NO 'Revisar campos'")
    r_po = _fake_result("PO-1", "F-12", "100.00", warn=False)
    r_po.document["po_reference"] = "PO-XYZ"
    check(ev.route_label(r_po.document) == "PO", "con OC referenciada -> ruta PO")
    r_conf = _fake_result("C-1", "F-13", "100.00", warn=False)
    r_conf.field_confidences = {"numero_factura": Decimal("0.8"), "importe_total": Decimal("0.6")}
    mconf = ev.aggregate_metrics([r_conf])
    check(mconf["confidence"] is not None and abs(mconf["confidence"] - 0.7) < 1e-9,
          "confianza promedio = media de confianzas informadas (0.7)")
    r_noconf = _fake_result("NC-1", "F-14", "100.00", warn=False)
    r_noconf.field_confidences = {}
    check(ev.aggregate_metrics([r_noconf])["confidence"] is None,
          "sin confianzas informadas -> confianza promedio None (no se inventa)")

    print("== add_document / add_error: tiempos por documento y estado de error ==")
    s3 = se.new_session()
    se.add_document(s3, r_ok, 0.5, file_hash="a" * 64, source="carga-manual")
    se.add_error(s3, "roto.pdf", "boom", 0.2)
    check(len(s3.results) == 1 and len(s3.errors) == 1, "un ok y un error en la sesion")
    check(s3.proc_seconds.get("OK-1") == 0.5 and abs(s3.processing_seconds - 0.7) < 1e-9,
          "tiempo por documento y total acumulado")
    check(s3.file_hashes.get("OK-1") == "a" * 64 and
          s3.sources.get("OK-1") == "carga-manual",
          "conserva hash y canal, nunca bytes del PDF")
    rows = ev._summary_rows(s3.results, s3.errors)
    need = {"archivo", "tipo", "proveedor", "número", "fecha", "vencimiento",
            "moneda", "total", "ruta PO/non-PO", "confianza", "estado"}
    check(need <= set(rows[0]), "la tabla tiene las columnas pedidas")
    check(rows[-1]["estado"] == "Error de procesamiento",
          "el archivo con error aparece con estado 'Error de procesamiento'")
    m3 = ev.aggregate_metrics(s3.results, s3.errors)
    check(m3["documents"] == 2 and m3["errors"] == 1,
          "'Documentos procesados' cuenta ok + errores")

    print()
    if failures:
        print(f"TRIAL SESSION ROJO: {len(failures)} fallas")
        return 1
    print("TRIAL SESSION VERDE: estado activo, audit sin PII, borrado y aislamiento (exit 0)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
