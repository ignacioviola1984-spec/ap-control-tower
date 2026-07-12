"""Regresión del workflow real: revisión humana + propuesta de pago."""

from __future__ import annotations

import sys
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


@dataclass
class FakeResult:
    doc_id: str
    document: dict
    confidence: Decimal = Decimal("0.90")
    warnings: list = field(default_factory=list)
    field_confidences: dict = field(default_factory=dict)
    engine: str = "google_document_ai_invoice_parser"
    pages: int = 1
    text_chars: int = 100


def invoice(doc_id: str, *, po=None, confidence="0.95") -> FakeResult:
    return FakeResult(doc_id, {
        "document_type": "invoice",
        "proveedor_nombre_comercial": "Proveedor SL",
        "numero_factura": doc_id,
        "fecha_emision": "2026-05-01",
        "fecha_vencimiento_calculada": "2026-05-31",
        "moneda": "EUR",
        "importe_total": "100.00",
        "po_reference": po,
    }, field_confidences={"numero_factura": Decimal(confidence),
                          "importe_total": Decimal("0.95")})


def must_raise(fn, text: str) -> None:
    try:
        fn()
    except ValueError as exc:
        assert text in str(exc), str(exc)
    else:
        raise AssertionError(f"se esperaba ValueError: {text}")


def main() -> int:
    from ap_control_tower.ui.trial.payment_approval import (
        _selected_doc_ids, payment_export_csv, payment_export_excel,
        payment_export_rows)
    from ap_control_tower.ui.trial import session as sess
    from ap_control_tower.ui.trial import workflow
    from ap_control_tower.ui.components import extraction_view as ev

    print("== Enrutamiento: OC ausente no implica revisión ==")
    clean = invoice("F-1", po=None)
    assert workflow.review_reasons(clean) == []
    assert workflow.approval_state(clean, {}, {})["status"] == "eligible"
    print("  PASS  factura non-PO limpia es elegible sin revisión")

    irrelevant = invoice("F-IRRELEVANTE")
    irrelevant.warnings = ["baja confianza en: cliente_nombre"]
    assert workflow.review_reasons(irrelevant) == []
    print("  PASS  campo informativo no deriva a revisión")

    print("== Baja confianza y campos críticos ==")
    low = invoice("F-2", confidence="0.50")
    low.warnings = ["baja confianza en: numero_factura"]
    assert any("baja confianza" in reason for reason in workflow.review_reasons(low))
    missing = invoice("F-3")
    missing.document["importe_total"] = None
    assert "importe_total" in workflow.missing_critical_fields(missing.document)
    assert workflow.approval_state(missing, {}, {})["status"] == "retained"
    print("  PASS  baja confianza y campos ausentes se retienen")

    print("== Proforma se deriva a revisión antes del gate de pago ==")
    proforma = invoice("P-1")
    proforma.document["document_type"] = "proforma_or_advance_request"
    state = workflow.approval_state(proforma, {}, {})
    assert state["status"] == "retained" and any(
        "no es una factura fiscal" in reason for reason in state["reasons"])
    proforma_queue = workflow.review_queue([proforma], {})
    assert len(proforma_queue) == 1 and proforma_queue[0]["pending"]
    print("  PASS  proforma retenida y visible automáticamente en revisión")

    print("== Duplicados ==")
    duplicate = invoice("F-1-COPY")
    duplicate.document["numero_factura"] = "F-1"
    ids = workflow.duplicate_doc_ids([clean, duplicate])
    assert ids == {"F-1", "F-1-COPY"}
    print("  PASS  mismo proveedor+número detectado")

    print("== Repetición técnica no es duplicado comercial ==")
    repeated_same_object = [clean, clean]
    assert workflow.unique_results(repeated_same_object) == [clean]
    assert workflow.duplicate_doc_ids(repeated_same_object) == set()
    assert len(workflow.review_queue(repeated_same_object, {})) == 0
    assert len(workflow.approval_rows(repeated_same_object, {}, {})) == 1
    print("  PASS  mismo doc_id repetido se colapsa y no genera falsa alerta")

    print("== Selección masiva de propuesta de pago ==")
    selection_labels = {"Factura 1": {"result": clean},
                        "Factura 2": {"result": duplicate}}
    assert _selected_doc_ids(selection_labels, [], True) == ["F-1", "F-1-COPY"]
    assert _selected_doc_ids(selection_labels, ["Factura 2"], False) == ["F-1-COPY"]
    assert _selected_doc_ids({}, [], True) == []
    print("  PASS  seleccionar todas no depende del valor visual del multiselect")

    print("== Decisión humana, maker-checker y no liberación ==")
    active = sess.new_session()
    active.results = [low, clean, proforma]
    sess.confirm_review(active, "F-2", "Ana Revisora", {
        field: str(low.document.get(field) or "") for field in workflow.EDITABLE_FIELDS
    }, "validado contra factura")
    assert active.review_decisions["F-2"]["status"] == "confirmed"
    assert workflow.approval_state(
        low, active.review_decisions, active.approval_decisions)["status"] == "eligible"
    must_raise(lambda: sess.decide_payment_proposal(
        active, ["F-2"], "Ana Revisora", "approved"), "Maker-checker")
    sess.decide_payment_proposal(active, ["F-2", "F-1"], "Bruno Aprobador",
                                 "approved", "lote propuesto")
    assert active.approval_decisions["F-2"]["status"] == "approved"
    last = active.audit.events[-1]
    assert last.action == "aprobada-para-propuesta-pago"
    assert last.evidence["no_libera_dinero"] is True
    must_raise(lambda: sess.decide_payment_proposal(
        active, ["P-1"], "Bruno Aprobador", "approved"), "no es elegible")
    must_raise(lambda: sess.approve_payment_exception(
        active, "P-1", "Ana Revisora", ""), "motivo")
    sess.approve_payment_exception(
        active, "P-1", "Ana Revisora", "anticipo validado por Finanzas")
    assert workflow.approval_state(
        proforma, active.review_decisions, active.approval_decisions)["status"] == "eligible"
    must_raise(lambda: sess.decide_payment_proposal(
        active, ["P-1"], "Ana Revisora", "approved"), "Maker-checker")
    sess.decide_payment_proposal(
        active, ["P-1"], "Bruno Aprobador", "approved", "anticipo en propuesta")
    assert active.approval_decisions["P-1"]["status"] == "approved"
    print("  PASS  excepción humana + maker-checker habilitan el anticipo")

    print("== Exportación del lote aprobado ==")
    approved_rows = [row for row in workflow.approval_rows(
        active.results, active.review_decisions, active.approval_decisions)
        if row["status"] == "approved"]
    export_rows = payment_export_rows(approved_rows)
    assert len(export_rows) == 3
    assert {row["estado"] for row in export_rows} == {"aprobada_para_propuesta"}
    csv_blob = payment_export_csv(approved_rows)
    assert b"beneficiario" in csv_blob and b"Bruno Aprobador" in csv_blob
    from io import BytesIO
    from openpyxl import load_workbook
    workbook = load_workbook(BytesIO(payment_export_excel(approved_rows)))
    assert workbook["Propuesta de pago"].max_row == 4
    print("  PASS  CSV y Excel contienen solo las tres aprobadas")

    print("== Retención/rechazo requiere motivo ==")
    must_raise(lambda: sess.retain_review(active, "F-1", "Ana", ""), "motivo")
    must_raise(lambda: sess.decide_payment_proposal(
        active, ["P-1"], "Bruno", "rejected", ""), "motivo")
    print("  PASS  decisiones negativas justificadas")

    print("== Invariante end-to-end: 1 revision -> 7 elegibles -> 7 aprobadas ==")
    seven = [invoice(f"INV-{index}") for index in range(1, 8)]
    seven[0].field_confidences["numero_factura"] = Decimal("0.50")
    seven[0].warnings = ["baja confianza en: numero_factura"]
    retained_proforma = invoice("PROFORMA-1")
    retained_proforma.document["document_type"] = "proforma_or_advance_request"
    run = sess.new_session()
    run.results = [*seven, retained_proforma]
    run.review_decisions["INV-2"] = {
        "status": "confirmed", "actor": "Revisor histórico",
        "timestamp": "2026-07-01T00:00:00Z"}
    queue = workflow.review_queue(run.results, run.review_decisions)
    assert {item["result"].doc_id for item in queue} == {"INV-1", "PROFORMA-1"}
    assert ev.aggregate_metrics(run.results)["with_warnings"] == len(queue)
    sess.confirm_review(run, "INV-1", "Revisora", {
        field: str(seven[0].document.get(field) or "")
        for field in workflow.EDITABLE_FIELDS
    }, "controlado")
    before = workflow.approval_rows(
        run.results, run.review_decisions, run.approval_decisions)
    assert sum(row["status"] == "eligible" for row in before) == 7
    assert sum(row["status"] == "approved" for row in before) == 0
    assert sum(row["status"] == "retained" for row in before) == 1
    sess.decide_payment_proposal(
        run, [item.doc_id for item in seven], "Aprobador", "approved", "propuesta")
    after = workflow.approval_rows(
        run.results, run.review_decisions, run.approval_decisions)
    assert sum(row["status"] == "eligible" for row in after) == 0
    assert sum(row["status"] == "approved" for row in after) == 7
    assert sum(row["status"] == "retained" for row in after) == 1
    assert all(event.evidence.get("no_libera_dinero") is True
               for event in run.audit.events
               if event.action == "aprobada-para-propuesta-pago")
    print("  PASS  estados 1/7/0/1 antes y 0/7/1 después, sin liberar dinero")

    print("== Retenida accionable y revisión por clasificación ==")
    sess.request_classification_review(
        run, "PROFORMA-1", "Analista", "validar si corresponde factura")
    manual_queue = workflow.review_queue(run.results, run.review_decisions)
    assert any(item["result"].doc_id == "PROFORMA-1" and item["pending"]
               for item in manual_queue)
    sess.decide_payment_proposal(
        run, ["PROFORMA-1"], "Aprobador", "excluded", "gestionar como anticipo")
    assert run.approval_decisions["PROFORMA-1"]["status"] == "excluded"
    assert run.audit.verify_chain()
    print("  PASS  clasificación solicitada y exclusión auditada")

    assert active.audit.verify_chain()
    print("\nTRIAL WORKFLOW VERDE: revisión, maker-checker, propuesta y audit trail")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
